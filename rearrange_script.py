from hungarian.hungarian import Hungarian
from openpyxl import Workbook, load_workbook
import numpy as np


def convert_to_array(sheet_row, num_columns):
    array = []
    for i in range(num_columns):
        val = sheet_row[i].internal_value
        if (val == None):
            array.append(0)
        array.append(sheet_row[i].internal_value)
    return array

def floating_equality(x,y):
    return np.abs(x - y ) < 0.001

def findNextStep(wb_in, name, row_index, timeStamp):
    row = wb_in.get_sheet_by_name(name)[row_index[name]]
    current_timeStamp = row[0].internal_value
    if (current_timeStamp == None):
        current_bodies.append([-1000] * num_columns)
        return
    elif floating_equality(current_timeStamp, timeStamp):
        current_bodies.append(convert_to_array(row, num_columns))
        row_index[name] = row_index[name] + 1
        return
    elif current_timeStamp < timeStamp:
        row_index[name] = row_index[name] + 1
        findNextStep(wb_in, name, row_index, timeStamp)
    else:
        current_bodies.append([-1000] * num_columns)
        return

# Use raw string, I'm tired of double backslashing
wb_in_location = r'G:\MR14012021\Session_20210114_32EMGs_Loomo_tracked_20210201095152\20210114_03_skeleton.xlsx'
wb_in = load_workbook(wb_in_location)

wb_in.remove_sheet(wb_in.get_sheet_by_name("Sheet"))

wb_out = Workbook()
wb_out.remove_sheet(wb_out.active)

matrix_dim = len(wb_in.sheetnames)

input_sheet_names = wb_in.sheetnames


output_sheets = []
output_lists = []
row_index = {}
i = 0
for sheet_name in input_sheet_names:
    row_index[sheet_name] = 1
    new_sheet = wb_out.create_sheet("Body " + str(i))
    output_sheets.append(new_sheet)
    output_lists.append([])
    i = i + 1

first_sheet = wb_in.get_sheet_by_name("Body 0")

is_first_row = True
is_first_data_row = False
previous_bodies = []

num_columns = 0

while True:
    timeStamp = first_sheet[row_index["Body 0"]][0].internal_value
    print(timeStamp)
    if is_first_row: # confirm that the spreadsheet looks as expected
        row = first_sheet[row_index["Body 0"]]
        if row[0].internal_value == "timeStamp(s)":
            print("workbook successfully loaded")
            is_first_row = False
            is_first_data_row = True
            num_columns = len(row)
        else:
            raise Exception("Unexpected file format")

        for key in row_index.keys():
            row_index[key] = row_index[key] + 1

    elif is_first_data_row: # setup previous frames rows
        is_first_data_row = False
        for name in input_sheet_names:
            array = convert_to_array(wb_in.get_sheet_by_name(name)[2], num_columns)
            previous_bodies.append(array)

        for key in row_index.keys():
            row_index[key] = row_index[key] + 1

    else: # standard execution
        if timeStamp == None: # finish
            break

        if floating_equality(timeStamp, 0.82):
            stopHere = True

        current_bodies = []
        for name in input_sheet_names:
            findNextStep(wb_in, name, row_index, timeStamp)

        deletion_tuples = []
        for j, current_body_j in enumerate(current_bodies):
            for i, current_body_i in enumerate(current_bodies):
                if np.linalg.norm(np.array(current_body_j) - np.array(current_body_i)) < 900:
                    deletion_tuples.append([j,i])
        
        for deletion_tuple in deletion_tuples:
            if (deletion_tuple[0] == deletion_tuple[1]):
                continue
            elif current_bodies[deletion_tuple[0]][0] == -1000:
                continue
            else:
                current_bodies[max(deletion_tuple[0], deletion_tuple[1])] = [-1000] * num_columns
                

        matrix = [[0 for i in range(len(previous_bodies))] for j in range(len(current_bodies))]
        for j, current_body in enumerate(current_bodies):
            for i, previous_body in enumerate(previous_bodies):
                distance = np.linalg.norm(np.array(current_body) - np.array(previous_body))
                matrix[j][i] = distance

        hung = Hungarian()
        hung.calculate(matrix)
        results = hung.get_results()

        current_bodies_reordered = [0] * len(current_bodies)
        for continuity_match in results:
            current_bodies_reordered[continuity_match[1]] = current_bodies[continuity_match[0]]
            output_sheets[continuity_match[1]].append(current_bodies[continuity_match[0]])
            output_lists[continuity_match[1]].append(current_bodies[continuity_match[0]])

        previous_bodies = current_bodies_reordered

wb_out.save(r'C:\Users\STIMO2-Admin\data\side_on_tracked_20210118120124\higher_sim_reject.xlsx')